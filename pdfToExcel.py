import os
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


df_all = pd.DataFrame()


def extract_date_from_filename(file_path):
    """
    从 PDF 文件名中提取日期（格式：YYYY_MM_DD）。
    """
    file_name = os.path.basename(file_path)  # 仅获取文件名
    match = re.search(r'(\d{4}_\d{2}_\d{2})', file_name)  # 匹配 YYYY_MM_DD 格式
    if match:
        return match.group(1)  # 返回匹配到的日期字符串
    return None  # 如果没有匹配到，返回 None


def convert_pdf_to_excel(pdf_path, excel_path):
    """
    提取 PDF 第一页中的表格，并保存为 Excel 文件。
    注意：此示例仅针对包含表格的 PDF，且只转换第一页数据。
    """
    # try:
    sheet_name = extract_date_from_filename(pdf_path)
    print(sheet_name)
    df = None
    global df_all
    with pdfplumber.open(pdf_path) as pdf:
        # 根据需要可以遍历所有页面
        page = pdf.pages[0]
        for pdf_table in page.find_tables():
            table = pdf_table.extract()
            if table is None:
                print(f"未在 {pdf_path} 中检测到表格数据。")
                return
            print(parse_table(table))

            df1 = pd.DataFrame(list(parse_table(table).items()), columns=['項目', sheet_name])
            df = pd.concat([df, df1], ignore_index=True)

        # print(df)
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        except FileNotFoundError:
            # 文件不存在时新建
            df.to_excel(excel_path, sheet_name=sheet_name, index=False)

        # 合并 DF1 和 DF2
        if sort_h:
            print("***********横向排列************")
            # 按“項目”合并
            if df_all.empty:
                df_all = df
            else:
                df_all = pd.merge(df_all, df, on="項目", how='outer')
            print(df_all)
        else:
            df_all = pd.concat([df_all, df], ignore_index=True)

        print(f"{pdf_path} 转换成功 -> {excel_path}")
    # except Exception as e:
    #     print(f"转换 {pdf_path} 时出错：{e}")


def parse_table(table):
    """
    将二维列表按两行一组（键行、值行）转换成字典，
    忽略空字符串或 None 的键和值。
    """
    result = {}
    # 每两行构成一组
    for i in range(0, len(table) - 1, 2):
        header = table[i]
        data = table[i+1]
        # 逐列处理
        for key, value in zip(header, data):
            if key and key.strip():  # 排除空的键
                # 若value存在且不为空，则将换行符去掉后保存
                if value and str(value).strip():
                    clean_key = key.replace("\n", "")
                    result[clean_key] = str(value).strip()
    return result


def batch_convert(folder_path, source_path):
    """
    遍历指定文件夹下所有 PDF 文件，并批量转换为 Excel。
    转换后的 Excel 文件保存在同一文件夹下，文件名与 PDF 相同，仅后缀不同。
    """
    # convert_pdf_to_excel(folder_path, source_path)
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(".pdf"):
            pdf_file = os.path.join(folder_path, filename)
            convert_pdf_to_excel(pdf_file, source_path)


def convert_time_to_minutes(time_str):
    if isinstance(time_str, str) and ":" in time_str:
        time_parts = time_str.split(":")
        if len(time_parts) == 3 or len(time_parts) == 2:  # 时:分:秒 时:分
            return int(time_parts[0]) * 60 + int(time_parts[1])
        elif len(time_parts) == 1:  # 时
            return int(time_parts[0])
    return time_str


def minutes_to_hours(minutes):
    hours, remainder = divmod(minutes, 60)
    return f"{int(hours):02}:{int(remainder):02}"


'''
提取PDF文件中的数据，转Excel表格
'''
if __name__ == "__main__":
    sort_h = True
    # 指定 PDF 文件所在的文件夹路径，请根据实际情况修改
    folder = r"C:\Users\yzok0\Downloads\salary"
    source_file = r"C:\Users\yzok0\Downloads\salary\result\result2.xlsx"
    batch_convert(folder, source_file)

    if sort_h:
        df_all = df_all.set_index("項目").T

        # # 预处理：去除数值列中的千位分隔符
        # for col in df_all.columns:
        #     # 1. 去除千位分隔符
        #     df_all[col] = df_all[col].replace({',': ''}, regex=True)
        #     # 2. 处理时间格式数据，转换为数字（例如小时和分钟转换为总分钟数）
        #     # 转换时间字段为分钟（如果是时间格式）
        #     df_all[col] = df_all[col].apply(lambda x: convert_time_to_minutes(x) if isinstance(x, str) and ":" in x else x)
        #     # 3. 转换为数值类型（去掉了千位分隔符和处理了时间字段后）
        #     df_all[col] = pd.to_numeric(df_all[col], errors='coerce')

        # 创建一个字典存放各列求和结果
        totals = {}

        for col in df_all.columns:
            # 取第一个非空样本数据用于判断
            sample = df_all[col].dropna().iloc[0] if not df_all[col].dropna().empty else None

            if sample is None:
                totals[col] = ''
            # 如果包含 ":" 则认为是时间列
            elif isinstance(sample, str) and ':' in sample:
                try:
                    # 转换整个列为 timedelta 类型
                    numeric_series = pd.to_numeric(
                        df_all[col].apply(
                            lambda x: convert_time_to_minutes(x) if isinstance(x, str) and ":" in x else x
                        )
                    )
                    total_td = minutes_to_hours(numeric_series.sum())
                    totals[col] = total_td
                    print(total_td)
                except Exception as e:
                    print(f"Exception : {e}")
                    totals[col] = ''
            # 如果是数字字符串（可能带逗号）
            elif isinstance(sample, str) and (',' in sample or sample.replace('.', '', 1).isdigit()):
                try:
                    # 去除逗号后转换为数值
                    numeric_series = pd.to_numeric(df_all[col].str.replace(',', ''), errors='coerce')
                    total_num = numeric_series.sum()
                    # 根据原始是否包含逗号来决定格式化方式
                    if ',' in sample:
                        totals[col] = f"{int(total_num):,}"
                    else:
                        totals[col] = str(total_num)
                except Exception as e:
                    totals[col] = ''
            # 如果已经是数值型数据
            elif pd.api.types.is_numeric_dtype(df_all[col]):
                total_num = df_all[col].sum()
                totals[col] = str(total_num)
            else:
                # 非求和列，留空或自定义
                totals[col] = ''

        # 将 totals 作为新行追加到 DataFrame 中，索引命名为“合计”
        df_all.loc['合计'] = pd.Series(totals)

        # # 对每列求和，并将结果作为新行添加
        # df_all.loc["合計"] = df_all.sum(axis=0)
        # 显示结果
        print(df_all)
        with pd.ExcelWriter(source_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_all.to_excel(writer, sheet_name='total', index=True, header=True)
    else:
        # 1. 去除千位分隔符
        df_all["值"] = df_all["值"].replace({',': ''}, regex=True)

        # 2. 处理时间格式数据，转换为数字（例如小时和分钟转换为总分钟数）
        # 转换时间字段为分钟（如果是时间格式）
        df_all["值"] = df_all["值"].apply(lambda x: convert_time_to_minutes(x) if isinstance(x, str) and ":" in x else x)

        # 3. 转换为数值类型（去掉了千位分隔符和处理了时间字段后）
        df_all["值"] = pd.to_numeric(df_all["值"], errors='coerce')

        # 4. 按項目分组并计算金额总和
        df_sum = df_all.groupby("項目")["值"].sum().reset_index()

        # 显示结果
        print(df_sum)
        with pd.ExcelWriter(source_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_sum.to_excel(writer, sheet_name='total', index=False)

    wb = load_workbook(source_file)

    # **遍历所有工作表**
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]  # 获取工作表

        # **调整列宽（根据最大内容自适应）**
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 9  # 适当增加列宽

        # **格式化标题行**
        header_font = Font(bold=True, color="FFFFFF")  # 加粗 + 白色字体
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 蓝色背景

        for cell in ws[1]:  # 第一行是标题行
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="right", vertical="center")  # 居中对齐

        # **格式化数据行**
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="right", vertical="center")  # 居中对齐

    # **保存 Excel**
    wb.save(source_file)
    wb.close()